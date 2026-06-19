# =====================================================================
# SYSTEMATIC ALPHA MULTI-ASSET SCANNER
# AUTHOR: QUANT RESEARCHER (AGE 15 PROTOTYPE V2.0 PRODUCTION)
# =====================================================================

# Step 1: Install and import required infrastructure
!pip install yfinance openpyxl --quiet
import pandas as pd
import yfinance as yf
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

print("Pipeline Initialized. Ingesting raw market arrays...")

# System Parameters & Risk Rules Configuration
PORTFOLIO_EQUITY = 100000.00   # Total account equity from ledger
MAX_RISK_PER_TRADE = 0.01      # 1% maximum portfolio risk rule
MAX_POSITION_ALLOCATION = 1/6  # Max 1/6th total portfolio value constraint

tickers = ["MSFT", "AMZN", "JPM", "XOM", "WMT", "UNH"]
sector_map = {
    "MSFT": "Technology", "AMZN": "Consumer Cyclical", "JPM": "Finance",
    "XOM": "Energy", "WMT": "Consumer Defensive", "UNH": "Health"
}

# Step 2: Live Market Filter - Validate S&P 500 Market Health
print("Evaluating S&P 500 Market Regime Filter...")
spy_data = yf.download("^GSPC", period="1y", interval="1d", progress=False)
spy_data["MA200"] = spy_data["Close"].rolling(200).mean()

spy_latest = spy_data.iloc[-1]
spy_close = spy_latest["Close"]
spy_ma200 = spy_latest["MA200"]
market_ok = spy_close > spy_ma200
print(f"S&P 500 Close: {spy_close:.2f} | 200MA: {spy_ma200:.2f} -> "
      f"{'RISK-ON' if market_ok else 'RISK-OFF'}")

# Step 3: Download and process isolated data loops
print("Downloading tracking asset universe...")
data = yf.download(tickers, period="1y", interval="1d", group_by='ticker', progress=False)

rows = []

for ticker in tickers:
    df_ticker = data[ticker].copy()
    df_ticker["Ticker"] = ticker
    df_ticker["Sector"] = sector_map[ticker]

    # Core structural moving average anchors
    df_ticker["MA200"] = df_ticker["Close"].rolling(200).mean()
    df_ticker["MA50"]  = df_ticker["Close"].rolling(50).mean()
    df_ticker["MA20"]  = df_ticker["Close"].rolling(20).mean()

    # Volatility Modeling (ATR 14)
    df_ticker["H-L"] = df_ticker["High"] - df_ticker["Low"]
    df_ticker["H-PC"] = (df_ticker["High"] - df_ticker["Close"].shift(1)).abs()
    df_ticker["L-PC"] = (df_ticker["Low"] - df_ticker["Close"].shift(1)).abs()
    df_ticker["TR"] = df_ticker[["H-L", "H-PC", "L-PC"]].max(axis=1)
    df_ticker["ATR"] = df_ticker["TR"].rolling(14).mean()

    # 30-Day Volume Moving Average
    df_ticker["VolAvg30"] = df_ticker["Volume"].rolling(30).mean()

    # Breakout Anchor (prior 20-day high)
    df_ticker["20d_high"] = df_ticker["High"].shift(1).rolling(20).max()

    rows.append(df_ticker)

# Step 4: Merge back into master research matrix
df = pd.concat(rows)
df.reset_index(inplace=True)

# Step 5: Vectorized Engine Logic Checks
df["MarketOK"] = market_ok
df["EarningsOK"] = True  # Placeholder for external earnings calendar data
df["TrendOK"] = df["Close"] > df["MA50"]

# Module A: Breakout
df["ModuleA"] = (
    (df["Close"] > df["20d_high"]) &
    (df["Volume"] > df["VolAvg30"]) &
    (df["TrendOK"]) &
    (df["EarningsOK"]) &
    (df["MarketOK"])
)

# Module B: Pullback
df["ModuleB"] = (
    (df["Close"] <= df["MA20"] + (2.0 * df["ATR"])) &
    (df["Close"] >= df["MA20"] - (2.0 * df["ATR"])) &
    (df["Volume"] < df["VolAvg30"]) &
    (df["TrendOK"]) &
    (df["EarningsOK"]) &
    (df["MarketOK"])
)

# Step 6: Extract Daily Signals and Calculate Capital Risks
latest_date = df["Date"].max()
today_df = df[df["Date"] == latest_date].copy()

final_report = []

for _, row in today_df.iterrows():
    ticker = row["Ticker"]
    close = row["Close"]
    atr = row["ATR"]

    if row["ModuleA"]:
        signal = "BUY_BREAKOUT"
        entry_type = "A"
    elif row["ModuleB"]:
        signal = "BUY_PULLBACK"
        entry_type = "B"
    else:
        signal = "NONE"
        entry_type = "-"

    stop_loss = close - (2.0 * atr) if signal != "NONE" else 0.0
    risk_per_share = close - stop_loss

    if signal != "NONE" and risk_per_share > 0:
        target_shares_by_risk = (PORTFOLIO_EQUITY * MAX_RISK_PER_TRADE) / risk_per_share
        target_shares_by_cap = (PORTFOLIO_EQUITY * MAX_POSITION_ALLOCATION) / close
        final_shares = int(min(target_shares_by_risk, target_shares_by_cap))
        final_allocation = final_shares * close
    else:
        stop_loss = "-"
        final_shares = "-"
        final_allocation = "-"

    final_report.append({
        "Ticker": ticker,
        "Sector": row["Sector"],
        "Entry Type": entry_type,
        "200-MA": "Above" if market_ok else "Below",
        "Entry Date": latest_date.strftime('%Y-%m-%d'),
        "Entry Price": round(close, 2),
        "Stop Loss": round(stop_loss, 2) if stop_loss != "-" else "-",
        "Position Size": final_shares,
        "Amount": round(final_allocation, 2) if final_allocation != "-" else "-",
        "Signal": signal
    })

output_df = pd.DataFrame(final_report)

# =====================================================================
# 🎨 EXCEL FORMATTING PIPELINE (FULLY FIXED)
# =====================================================================
file_name = "todays_signals.xlsx"

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    output_df.to_excel(writer, index=False, sheet_name='Daily Report')

    workbook = writer.book
    worksheet = writer.sheets['Daily Report']

    # Correct gridline access
    worksheet.sheet_view.showGridLines = True

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    signal_a_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    signal_b_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=10, color="333333")
    font_sig_a = Font(name="Segoe UI", size=10, bold=True, color="375623")
    font_sig_b = Font(name="Segoe UI", size=10, bold=True, color="7F6000")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side,
                         top=thin_border_side, bottom=thin_border_side)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Format Headers (row 1)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Format Data Rows
    for r_idx in range(2, len(output_df) + 2):
        worksheet.row_dimensions[r_idx].height = 20
        sig_val = worksheet.cell(row=r_idx, column=10).value
        is_signal = sig_val != "NONE"

        for c_idx in range(1, 11):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.font = body_font
            cell.border = thin_border

            # Zebra striping for non-signal rows
            if r_idx % 2 == 0 and not is_signal:
                cell.fill = zebra_fill

            # Alignment rules
            # 1: Ticker, 3: Entry Type, 4: 200-MA, 5: Entry Date centered
            if c_idx in [1, 3, 4, 5]:
                cell.alignment = center_align
            elif c_idx == 2:  # Sector
                cell.alignment = left_align
            else:
                cell.alignment = right_align if c_idx != 10 else center_align

            # Number formatting
            val = cell.value
            if val != "-":
                # 6: Entry Price, 7: Stop Loss, 9: Amount
                if c_idx in [6, 7, 9]:
                    cell.number_format = '$#,##0.00'
                elif c_idx == 8:  # Position Size
                    cell.number_format = '#,##0'

            # Conditional signal highlighting
            if c_idx == 10:
                if val == "BUY_BREAKOUT":
                    cell.fill = signal_a_fill
                    cell.font = font_sig_a
                elif val == "BUY_PULLBACK":
                    cell.fill = signal_b_fill
                    cell.font = font_sig_b

    # Auto-adjust column sizes
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = max(max_len + 4, 13)

print(f"\n🚀 Production Success! Engineered '{file_name}' grid snapshot for {latest_date.date()}")
print(output_df.to_string(index=False))
