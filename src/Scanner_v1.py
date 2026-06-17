# =====================================================================
# SYSTEMATIC ALPHA MULTI-ASSET SCANNER
# AUTHOR: QUANT RESEARCHER (AGE 15 PROTOTYPE V1.0)
# =====================================================================

# Step 1: Install and import required infrastructure
!pip install yfinance openpyxl
import pandas as pd
import yfinance as yf
from datetime import datetime

print("Pipeline Initialized. Ingesting raw market arrays...")

# Step 2: Define your core asset tracking universe
tickers = ["MSFT", "AMZN", "JPM", "XOM", "WMT", "UNH"]

# Download 1 year of historical data to accurately calculate 200-day lines
data = yf.download(tickers, period="1y", interval="1d", group_by='ticker')

rows = []

# Step 3: Process mathematical metrics inside isolated loops to prevent cross-ticker data bleed
for ticker in tickers:
    df_ticker = data[ticker].copy()
    df_ticker["Ticker"] = ticker

    # Core structural moving average anchors
    df_ticker["MA200"] = df_ticker["Close"].rolling(200).mean()
    df_ticker["MA50"]  = df_ticker["Close"].rolling(50).mean()
    df_ticker["MA20"]  = df_ticker["Close"].rolling(20).mean()

    # Institutional True Range (TR) & Volatility Modeling (ATR)
    df_ticker["H-L"] = df_ticker["High"] - df_ticker["Low"]
    df_ticker["H-PC"] = (df_ticker["High"] - df_ticker["Close"].shift(1)).abs()
    df_ticker["L-PC"] = (df_ticker["Low"] - df_ticker["Close"].shift(1)).abs()
    df_ticker["TR"] = df_ticker[["H-L", "H-PC", "L-PC"]].max(axis=1)
    df_ticker["ATR"] = df_ticker["TR"].rolling(14).mean()

    # 30-Day Volume Moving Average
    df_ticker["VolAvg30"] = df_ticker["Volume"].rolling(30).mean()

    # Prior 20-Day Breakout Anchor (Shifted by 1 to exclude today's high)
    df_ticker["20d_high"] = df_ticker["High"].shift(1).rolling(20).max()

    rows.append(df_ticker)

# Step 4: Merge isolated streams back into a master research matrix
df = pd.concat(rows)
df.reset_index(inplace=True)

# Step 5: Initialize Macro Gatekeepers and Structural Placeholders
market_ok = True      # Core breadth index placeholder
df["EarningsOK"] = True  # Corporate actions calendar placeholder

# Step 6: Define Core Vectorized System Checklists
df["TrendOK"] = df["Close"] > df["MA50"]
df["VolumeOK"] = df["Volume"] > df["VolAvg30"]

# Module A: Momentum Breakout Engine Logic Matrix
df["ModuleA"] = (
    (df["Close"] > df["20d_high"]) &
    (df["TrendOK"]) &
    (df["VolumeOK"]) &
    (df["EarningsOK"]) &
    (market_ok)
)

# Module B: Structural Pullback Engine Logic Matrix (2.0x ATR boundary around the 20MA)
df["ModuleB"] = (
    (df["Close"] < df["MA20"] + (2.0 * df["ATR"])) &
    (df["Close"] > df["MA20"] - (2.0 * df["ATR"])) &
    (df["Volume"] < df["VolAvg30"]) & # Rule: Volume must contract on pullbacks
    (df["TrendOK"]) &
    (df["EarningsOK"]) &
    (market_ok)
)

# Step 7: Route Logic Output Signals
def final_signal(row):
    if row["ModuleA"]: return "BUY_BREAKOUT"
    if row["ModuleB"]: return "BUY_PULLBACK"
    return "NONE"

df["Signal"] = df.apply(final_signal, axis=1)

# =====================================================================
# 🎯 THE DATA SLICE INTERCEPTION: ISOLATE TODAY'S SIGNALS
# =====================================================================
latest_date = df["Date"].max()
today_signals = df[df["Date"] == latest_date].copy()

# Keep only relevant operational columns for the daily layout sheet
columns_to_keep = ["Date", "Ticker", "Close", "Volume", "MA50", "ATR", "Signal"]
today_signals_clean = today_signals[columns_to_keep]

# Format the date text column nicely
today_signals_clean["Date"] = today_signals_clean["Date"].dt.strftime('%Y-%m-%d')

# =====================================================================
# 🎨 EXCEL FORMATTING PIPELINE (INSTITUTIONAL PRESENTATION)
# =====================================================================
file_name = "todays_signals.xlsx"

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    today_signals_clean.to_excel(writer, index=False, sheet_name='Daily Report')

    # Grab the worksheet layout details
    workbook = writer.book
    worksheet = writer.sheets['Daily Report']

    # Import stylers for layout construction
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # Palette definition: Charcoal headers, Light Mint highlights for active signals
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    signal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, bold=False)
    signal_font = Font(name="Calibri", size=11, bold=True, color="375623")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Format Headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Format Data Rows
    for row in worksheet.iter_rows(min_row=2, max_row=len(today_signals_clean)+1):
        row[0].alignment = center_align # Date
        row[1].alignment = center_align # Ticker
        row[2].number_format = '$#,##0.00' # Close Price
        row[2].alignment = right_align
        row[3].number_format = '#,##0'    # Volume
        row[3].alignment = right_align
        row[4].number_format = '$#,##0.00' # MA50
        row[4].alignment = right_align
        row[5].number_format = '$#,##0.00' # ATR
        row[5].alignment = right_align
        row[6].alignment = center_align # Signal

        # Color-code active execution triggers
        if row[6].value in ["BUY_BREAKOUT", "BUY_PULLBACK"]:
            row[6].fill = signal_fill
            row[6].font = signal_font

    # Automatically widen the Excel columns to match content width
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

print(f"\n🚀 Success! Engineered '{file_name}' grid snapshot for {latest_date.date()}")
print(today_signals_clean.to_string(index=False))
